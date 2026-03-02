"""
Consolidador — Gerador de Relatório Excel.

Recebe dados consolidados e gera Excel com 6 abas:
1. Resumo — visão macro por conta
2. Alocação — distribuição por estratégia e corretora
3. Posição Detalhada — todos os ativos de todas as contas
4. Rentabilidade — histórico mensal por conta
5. Evolução Patrimonial — por conta
6. Movimentações — lista unificada
"""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ============================================================================
# ESTILOS
# ============================================================================

# Cores
AZUL_ESCURO = "1B2A4A"
AZUL_MEDIO = "2E4057"
AZUL_CLARO = "E8EDF2"
CINZA_CLARO = "F5F5F5"
BRANCO = "FFFFFF"
VERDE = "27AE60"
VERMELHO = "E74C3C"

# Fontes
FONT_TITULO = Font(name="Calibri", size=14, bold=True, color=BRANCO)
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color=BRANCO)
FONT_SUBHEADER = Font(name="Calibri", size=11, bold=True, color=AZUL_ESCURO)
FONT_NORMAL = Font(name="Calibri", size=10)
FONT_BOLD = Font(name="Calibri", size=10, bold=True)
FONT_TOTAL = Font(name="Calibri", size=11, bold=True, color=AZUL_ESCURO)

# Fills
FILL_HEADER = PatternFill("solid", fgColor=AZUL_ESCURO)
FILL_SUBHEADER = PatternFill("solid", fgColor=AZUL_CLARO)
FILL_ALTERNATING = PatternFill("solid", fgColor=CINZA_CLARO)
FILL_TOTAL = PatternFill("solid", fgColor=AZUL_CLARO)

# Alinhamentos
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Bordas
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# Formatos numéricos brasileiros para openpyxl
FMT_CURRENCY = '#,##0.00'
FMT_PCT = '0.00"%"'
FMT_NUMBER = '#,##0.00'


def _write_header_row(ws, row, headers, col_start=1):
    """Escreve linha de cabeçalho com estilo."""
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def _write_data_row(ws, row, values, col_start=1, bold=False, is_total=False):
    """Escreve linha de dados com formatação."""
    font = FONT_TOTAL if is_total else (FONT_BOLD if bold else FONT_NORMAL)
    fill = FILL_TOTAL if is_total else (FILL_ALTERNATING if row % 2 == 0 else None)
    
    for i, value in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i, value=value)
        cell.font = font
        if fill:
            cell.fill = fill
        cell.border = THIN_BORDER
        
        # Formatação por tipo
        if isinstance(value, float):
            cell.alignment = ALIGN_RIGHT
            cell.number_format = FMT_NUMBER
        elif isinstance(value, (int,)):
            cell.alignment = ALIGN_RIGHT
        else:
            cell.alignment = ALIGN_LEFT


def _format_currency_cell(cell):
    """Aplica formatação de moeda a uma célula."""
    cell.number_format = FMT_CURRENCY
    cell.alignment = ALIGN_RIGHT


def _format_pct_cell(cell):
    """Aplica formatação de percentual a uma célula."""
    cell.number_format = FMT_PCT
    cell.alignment = ALIGN_RIGHT


def _auto_width(ws, min_width=8, max_width=40):
    """Ajusta largura das colunas automaticamente."""
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                length = len(str(cell.value))
                max_len = max(max_len, min(length + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def _write_section_title(ws, row, title, num_cols=10):
    """Escreve título de seção."""
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = FONT_SUBHEADER
    cell.fill = FILL_SUBHEADER
    for i in range(1, num_cols + 1):
        ws.cell(row=row, column=i).fill = FILL_SUBHEADER
        ws.cell(row=row, column=i).border = THIN_BORDER


# ============================================================================
# ABA 1: RESUMO
# ============================================================================

def _write_resumo(ws, data):
    """Aba Resumo — visão macro por conta + benchmarks."""
    ws.title = "Resumo"
    
    # Título
    ws.merge_cells("A1:I1")
    cell = ws.cell(row=1, column=1, value=f"Consolidação — {data['cliente']}")
    cell.font = FONT_TITULO
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER

    ws.merge_cells("A2:I2")
    cell = ws.cell(row=2, column=1, value=f"Data de Referência: {data.get('data_referencia', '')}")
    cell.font = Font(name="Calibri", size=10, color=BRANCO)
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    
    # Tabela de contas
    row = 4
    headers = ["Corretora", "Conta", "Patrimônio Bruto",
               "Rent. Mês (%)", "%CDI Mês", "Rent. Ano (%)", "%CDI Ano",
               "Ganho Mês (R$)", "Ganho Ano (R$)"]
    _write_header_row(ws, row, headers)

    row = 5
    for conta in data.get("contas", []):
        values = [
            conta.get("corretora", ""),
            conta.get("conta", ""),
            conta.get("patrimonio_bruto", 0),
            conta.get("rentabilidade_mes_pct"),
            conta.get("pct_cdi_mes"),
            conta.get("rentabilidade_ano_pct"),
            conta.get("pct_cdi_ano"),
            conta.get("ganho_mes_rs"),
            conta.get("ganho_ano_rs"),
        ]
        _write_data_row(ws, row, values)

        # Formatar colunas numéricas
        _format_currency_cell(ws.cell(row=row, column=3))
        for col in [4, 5, 6, 7]:
            _format_pct_cell(ws.cell(row=row, column=col))
        for col in [8, 9]:
            _format_currency_cell(ws.cell(row=row, column=col))

        row += 1

    # Linha Total
    total_values = [
        "TOTAL", "",
        data.get("patrimonio_total_consolidado", 0),
        None, None, None, None,
        sum(c.get("ganho_mes_rs", 0) or 0 for c in data.get("contas", [])),
        sum(c.get("ganho_ano_rs", 0) or 0 for c in data.get("contas", [])),
    ]
    _write_data_row(ws, row, total_values, is_total=True)
    _format_currency_cell(ws.cell(row=row, column=3))
    _format_currency_cell(ws.cell(row=row, column=8))
    _format_currency_cell(ws.cell(row=row, column=9))
    
    # Benchmarks
    row += 2
    benchmarks = data.get("benchmarks", {})
    if benchmarks:
        _write_section_title(ws, row, "Benchmarks do Período", 5)
        row += 1
        _write_header_row(ws, row, ["Índice", "Mês (%)", "Ano (%)", "12M (%)", "24M (%)"])
        row += 1
        
        for idx_name in ["cdi", "ibovespa", "ipca", "dolar"]:
            idx_data = benchmarks.get(idx_name, {})
            if idx_data:
                values = [
                    idx_name.upper(),
                    idx_data.get("mes"),
                    idx_data.get("ano"),
                    idx_data.get("12m"),
                    idx_data.get("24m"),
                ]
                _write_data_row(ws, row, values)
                for col in [2, 3, 4, 5]:
                    _format_pct_cell(ws.cell(row=row, column=col))
                row += 1
    
    _auto_width(ws)


# ============================================================================
# ABA 2: ALOCAÇÃO
# ============================================================================

def _write_alocacao(ws, data):
    """Aba Alocação — distribuição por estratégia, corretora, tipo."""
    ws.title = "Alocação"
    
    row = 1
    
    # Tabela 1: Por Estratégia
    _write_section_title(ws, row, "Alocação por Estratégia", 3)
    row += 1
    _write_header_row(ws, row, ["Estratégia", "Saldo Bruto (R$)", "% Total"])
    row += 1
    
    for item in data.get("alocacao_por_estrategia", []):
        values = [item["estrategia"], item["saldo_bruto"], item["pct_total"]]
        _write_data_row(ws, row, values)
        _format_currency_cell(ws.cell(row=row, column=2))
        _format_pct_cell(ws.cell(row=row, column=3))
        row += 1
    
    # Total
    _write_data_row(ws, row, [
        "TOTAL",
        data.get("patrimonio_total_consolidado", 0),
        100.0,
    ], is_total=True)
    _format_currency_cell(ws.cell(row=row, column=2))
    _format_pct_cell(ws.cell(row=row, column=3))
    row += 2
    
    # Tabela 2: Por Corretora
    _write_section_title(ws, row, "Alocação por Corretora", 3)
    row += 1
    _write_header_row(ws, row, ["Corretora (Conta)", "Saldo Bruto (R$)", "% Total"])
    row += 1
    
    for item in data.get("alocacao_por_corretora", []):
        values = [item["corretora"], item["saldo_bruto"], item["pct_total"]]
        _write_data_row(ws, row, values)
        _format_currency_cell(ws.cell(row=row, column=2))
        _format_pct_cell(ws.cell(row=row, column=3))
        row += 1
    
    _auto_width(ws)


# ============================================================================
# ABA 3: POSIÇÃO DETALHADA
# ============================================================================

def _write_posicao(ws, data):
    """Aba Posição Detalhada — todos os ativos, todas as contas."""
    ws.title = "Posição Detalhada"
    
    headers = [
        "Corretora", "Conta", "Estratégia", "Ativo",
        "Saldo Bruto (R$)", "% Total", "Rent. Mês (%)",
        "%CDI Mês", "Rent. Ano (%)", "%CDI Ano"
    ]
    _write_header_row(ws, 1, headers)

    row = 2
    for ativo in data.get("ativos_consolidados", []):
        values = [
            ativo.get("corretora", ""),
            ativo.get("conta", ""),
            ativo.get("estrategia_normalizada", ativo.get("estrategia", "")),
            ativo.get("nome_original", ""),
            ativo.get("saldo_bruto", 0),
            ativo.get("pct_total_consolidado", 0),
            ativo.get("rent_mes_pct"),
            ativo.get("pct_cdi_mes"),
            ativo.get("rent_ano_pct"),
            ativo.get("pct_cdi_ano"),
        ]
        _write_data_row(ws, row, values)
        _format_currency_cell(ws.cell(row=row, column=5))
        for col in [6, 7, 8, 9, 10]:
            _format_pct_cell(ws.cell(row=row, column=col))
        row += 1

    # Total
    total_saldo = sum(a.get("saldo_bruto", 0) for a in data.get("ativos_consolidados", []))
    _write_data_row(ws, row, [
        "TOTAL", "", "", "",
        total_saldo, 100.0,
        None, None, None, None,
    ], is_total=True)
    _format_currency_cell(ws.cell(row=row, column=5))
    _format_pct_cell(ws.cell(row=row, column=6))

    _auto_width(ws)
    ws.column_dimensions["D"].width = 50  # Ativo


# ============================================================================
# ABA 4: RENTABILIDADE
# ============================================================================

def _write_rentabilidade(ws, data):
    """Aba Rentabilidade — histórico mensal por conta."""
    ws.title = "Rentabilidade"
    
    MESES = ["jan", "fev", "mar", "abr", "mai", "jun",
             "jul", "ago", "set", "out", "nov", "dez"]
    
    row = 1
    
    for conta_data in data.get("rentabilidade_por_conta", []):
        corretora = conta_data.get("corretora", "?")
        conta = conta_data.get("conta", "?")
        
        # Título da conta
        _write_section_title(ws, row, f"{corretora} — Conta {conta}", 16)
        row += 1
        
        # Headers: Ano | Jan | Fev | ... | Dez | Ano (%) | Acum (%)
        headers = ["Ano", "Tipo"] + [m.capitalize() for m in MESES] + ["Ano (%)", "Acum. (%)"]
        _write_header_row(ws, row, headers)
        row += 1
        
        for year_data in conta_data.get("rentabilidade_historica_mensal", []):
            ano = year_data.get("ano", "")
            meses_data = year_data.get("meses", {})
            
            # Linha Portfólio %
            values = [ano, "Portfólio %"]
            for mes in MESES:
                mes_info = meses_data.get(mes, {})
                values.append(mes_info.get("portfolio_pct") if mes_info else None)
            values.append(year_data.get("ano_pct"))
            values.append(year_data.get("acumulada_pct"))
            _write_data_row(ws, row, values)
            for col in range(3, 17):
                _format_pct_cell(ws.cell(row=row, column=col))
            row += 1
            
            # Linha %CDI
            values = [ano, "%CDI"]
            for mes in MESES:
                mes_info = meses_data.get(mes, {})
                values.append(mes_info.get("pct_cdi") if mes_info else None)
            values.append(None)
            values.append(None)
            _write_data_row(ws, row, values)
            for col in range(3, 17):
                _format_pct_cell(ws.cell(row=row, column=col))
            row += 1
        
        row += 1  # Espaço entre contas
    
    _auto_width(ws)


# ============================================================================
# ABA 5: EVOLUÇÃO PATRIMONIAL
# ============================================================================

def _write_evolucao(ws, data):
    """Aba Evolução Patrimonial — por conta."""
    ws.title = "Evolução Patrimonial"
    
    row = 1
    
    for conta_data in data.get("evolucao_por_conta", []):
        corretora = conta_data.get("corretora", "?")
        conta = conta_data.get("conta", "?")
        
        # Título
        _write_section_title(ws, row, f"{corretora} — Conta {conta}", 9)
        row += 1
        
        headers = ["Data", "Patrimônio Inicial", "Movimentações", "IR", "IOF",
                   "Patrimônio Final", "Ganho Financeiro", "Rent. (%)", "%CDI"]
        _write_header_row(ws, row, headers)
        row += 1
        
        for ep in conta_data.get("evolucao_patrimonial", []):
            values = [
                ep.get("data", ""),
                ep.get("patrimonio_inicial", 0),
                ep.get("movimentacoes", 0),
                ep.get("ir", 0),
                ep.get("iof", 0),
                ep.get("patrimonio_final", 0),
                ep.get("ganho_financeiro", 0),
                ep.get("rentabilidade_pct"),
                ep.get("pct_cdi"),
            ]
            _write_data_row(ws, row, values)
            for col in [2, 3, 4, 5, 6, 7]:
                _format_currency_cell(ws.cell(row=row, column=col))
            for col in [8, 9]:
                _format_pct_cell(ws.cell(row=row, column=col))
            row += 1
        
        row += 1  # Espaço entre contas
    
    _auto_width(ws)


# ============================================================================
# ABA 6: MOVIMENTAÇÕES
# ============================================================================

def _write_movimentacoes(ws, data):
    """Aba Movimentações — lista unificada ordenada por data."""
    ws.title = "Movimentações"
    
    headers = ["Data Mov", "Data Liq", "Corretora", "Conta", "Histórico", "Valor (R$)", "Saldo (R$)"]
    _write_header_row(ws, 1, headers)
    
    row = 2
    for mov in data.get("movimentacoes_unificadas", []):
        values = [
            mov.get("data_mov", ""),
            mov.get("data_liq", ""),
            mov.get("corretora", ""),
            mov.get("conta", ""),
            mov.get("historico", ""),
            mov.get("valor", 0),
            mov.get("saldo"),
        ]
        _write_data_row(ws, row, values)
        _format_currency_cell(ws.cell(row=row, column=6))
        if mov.get("saldo") is not None:
            _format_currency_cell(ws.cell(row=row, column=7))
        row += 1
    
    _auto_width(ws)
    ws.column_dimensions["E"].width = 50  # Histórico mais largo


# ============================================================================
# FUNÇÃO PRINCIPAL
# ============================================================================

def generate_report(consolidated: dict, output_path: str):
    """
    Gera relatório Excel consolidado com 6 abas.
    
    Args:
        consolidated: Dict com dados consolidados (output do consolidator).
        output_path: Caminho para salvar o arquivo .xlsx.
    """
    logger.info(f"Gerando relatório Excel: {output_path}")
    
    wb = Workbook()
    
    # Aba 1: Resumo (usa a aba padrão)
    _write_resumo(wb.active, consolidated)
    
    # Aba 2: Alocação
    ws2 = wb.create_sheet()
    _write_alocacao(ws2, consolidated)
    
    # Aba 3: Posição Detalhada
    ws3 = wb.create_sheet()
    _write_posicao(ws3, consolidated)
    
    # Aba 4: Rentabilidade
    ws4 = wb.create_sheet()
    _write_rentabilidade(ws4, consolidated)
    
    # Aba 5: Evolução Patrimonial
    ws5 = wb.create_sheet()
    _write_evolucao(ws5, consolidated)
    
    # Aba 6: Movimentações
    ws6 = wb.create_sheet()
    _write_movimentacoes(ws6, consolidated)
    
    # Salvar
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    n_ativos = len(consolidated.get("ativos_consolidados", []))
    n_mov = len(consolidated.get("movimentacoes_unificadas", []))
    logger.info(
        f"Relatório salvo: {output_path} | "
        f"6 abas | {n_ativos} ativos | {n_mov} movimentações"
    )
